import openpyxl
book = openpyxl.load_workbook("D:\\WorkSpace\\ExcelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
#print(cell.value)
sheet.cell(row=2, column=2).value = "Neeraj"
#print(sheet.cell(row=2, column=2).value)
#print(sheet.max_column)
#print(sheet.max_row)
#print(sheet['A2'].value)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)
